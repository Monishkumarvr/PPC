"""
GRINDING OPTIMIZATION SYSTEM - COMPLETE VERSION
Refactored for Streamlit Integration
"""

import math
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
from collections import defaultdict
import io

import pandas as pd
import pulp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================================
# GLOBAL CONFIG
# ============================================================================

class GrindingConfig:
    """Configuration for grinding optimization"""
    
    def __init__(self, 
                 master_data: Optional[Dict[str, pd.DataFrame]] = None,
                 grinding_resources: int = 35,
                 hours_per_day: float = 8.0,
                 line_oee: float = 0.90,
                 lead_time_days: int = 2,
                 shortage_penalty: float = 100000000.0,
                 lateness_penalty: float = 50.0,
                 production_lateness_penalty: float = 1000.0,
                 leadtime_required_days: int = 14,
                 leadtime_penalty: float = 25.0,
                 early_penalty: float = 0.0,
                 solver_timeout: int = 600):
        
        # Resource capacity
        self.GRINDING_RESOURCES = grinding_resources
        self.HOURS_PER_DAY = hours_per_day
        self.LINE_OEE = line_oee
        self.EFFECTIVE_DAY_MIN = hours_per_day * 60.0 * line_oee
        
        # Inter-stage lead time
        self.LEAD_TIME_DAYS = lead_time_days
        
        # Penalty weights
        self.SHORTAGE_PENALTY = shortage_penalty
        self.LATENESS_PENALTY_PER_DAY = lateness_penalty
        self.PRODUCTION_LATENESS_PENALTY = production_lateness_penalty
        self.LEADTIME_REQUIRED_DAYS = leadtime_required_days
        self.LEADTIME_PENALTY_PER_DAY = leadtime_penalty
        self.EARLY_PRODUCTION_PENALTY = early_penalty
        
        # Solver settings
        self.SOLVER_TIME_LIMIT = solver_timeout
        self.SOLVER_GAP_REL = 0.005
        self.SOLVER_THREADS = 4 # Reduced for cloud environments
        
        # Capacities
        self.capacity_by_resource_code: Dict[str, float] = {}
        self.resource_meta: Dict[str, Dict[str, object]] = {}
        self.total_grinding_capacity_per_day = 0.0

        if master_data:
            self.load_machine_constraints(master_data)

        # Fallback if machine constraints not available
        if not self.capacity_by_resource_code:
            self.total_grinding_capacity_per_day = (
                self.EFFECTIVE_DAY_MIN * self.GRINDING_RESOURCES
            )

    def load_machine_constraints(self, master_data: Dict[str, pd.DataFrame]) -> None:
        """Load grinding capacities from Master Data"""
        try:
            if "Machine Constraints" not in master_data:
                return

            
            # Create a copy to avoid modifying the original dataframe in session state
            mc = master_data["Machine Constraints"].copy()
            # Normalize column names
            mc.columns = [str(c).strip() for c in mc.columns]

            if "Operation Name" not in mc.columns:
                return

            grind = mc[mc["Operation Name"].astype(str).str.contains("grind", case=False, na=False)].copy()
            if grind.empty:
                return

            for _, r in grind.iterrows():
                rc = str(r.get("Resource Code", "")).strip()
                if not rc or rc.lower() == "nan":
                    continue

                n = float(r.get("No Of Resource", 0) or 0)
                hrs = float(r.get("Available Hours per Day", 0) or 0)
                if n <= 0 or hrs <= 0:
                    continue

                cap = hrs * 60.0 * self.LINE_OEE * n
                self.capacity_by_resource_code[rc] = cap
                self.resource_meta[rc] = {
                    "production_unit": r.get("Production unit", ""),
                    "resource_name": r.get("Resource Name", ""),
                    "operation_name": r.get("Operation Name", ""),
                    "no_resources": n,
                    "hours_per_day": hrs,
                    "oee": self.LINE_OEE,
                    "capacity_min_per_day": cap,
                }

            # Refresh total capacity
            if self.capacity_by_resource_code:
                self.total_grinding_capacity_per_day = sum(self.capacity_by_resource_code.values())

        except Exception as e:
            print(f"[!] WARNING: Could not load grinding capacity: {e}")


# ============================================================================
# WIP AND CASTING INTEGRATION
# ============================================================================

class CastingScheduleReader:
    """Reads casting schedule DataFrame and extracts CS availability"""
    
    def __init__(self, schedule_df: pd.DataFrame, config: GrindingConfig):
        self.schedule_df = schedule_df
        self.config = config
    
    def load_cs_availability(
        self, 
        start_date: datetime, 
        end_date: datetime
    ) -> Dict[str, Dict[int, float]]:
        
        try:
            if self.schedule_df is None or self.schedule_df.empty:
                return {}

            # Generate working days
            all_days = pd.date_range(start_date, end_date, freq="D")
            days = [d for d in all_days if d.weekday() < 6]
            day_to_index = {d.date(): i for i, d in enumerate(days)}
            
            # Map columns
            df = self.schedule_df.copy()
            
            # Normalize columns if needed (assuming standard output format)
            # Standard columns: 'FG Code', 'Date', 'Units'
            
            # Track CS production by day
            cs_production_by_day = defaultdict(lambda: defaultdict(float))
            LEAD_TIME_DAYS = self.config.LEAD_TIME_DAYS
            
            for _, row in df.iterrows():
                fg_code = str(row.get("FG Code", "")).strip()
                if not fg_code: continue
                
                try:
                    cast_date = pd.to_datetime(row.get("Date")).date()
                except:
                    continue
                    
                if cast_date not in day_to_index:
                    continue
                
                cast_day_idx = day_to_index[cast_date]
                units = float(row.get("Units", 0) or 0)
                
                if units > 0:
                    # Apply lead time
                    available_day_idx = cast_day_idx + LEAD_TIME_DAYS
                    
                    if available_day_idx < len(days):
                        cs_production_by_day[fg_code][available_day_idx] += units
            
            # Convert to cumulative
            cs_cumulative = {}
            for fg_code, daily_prod in cs_production_by_day.items():
                cs_cumulative[fg_code] = {}
                cumulative = 0.0
                for day_idx in range(len(days)):
                    cumulative += daily_prod.get(day_idx, 0)
                    cs_cumulative[fg_code][day_idx] = cumulative
            
            return cs_cumulative
            
        except Exception as e:
            print(f"WARNING: Error processing casting schedule: {e}")
            return {}


class WIPLoader:
    """Loads WIP inventory from master data dict"""
    
    def __init__(self, master_data: Dict[str, pd.DataFrame]):
        self.master_data = master_data
    
    def load_wip_inventory(self) -> Dict[str, Dict[str, float]]:
        try:
            if "Stage WIP" not in self.master_data:
                return {}
            
            wip_df = self.master_data["Stage WIP"]
            wip_inventory = {}
            
            for _, row in wip_df.iterrows():
                casting_item = str(row.get("CastingItem", "")).strip()
                if not casting_item:
                    continue
                
                # Handle CS1- and CS- prefixes (Logic ported from updated optimization_engine)
                if casting_item.upper().startswith("CS1-"):
                    fg_code = casting_item[4:]
                elif casting_item.upper().startswith("CS-"):
                    fg_code = casting_item[3:]
                else:
                    fg_code = casting_item
                
                cs_qty = float(row.get("CS", 0) or 0)
                gr_qty = float(row.get("GR", 0) or 0)
                mc_qty = float(row.get("MC", 0) or 0)
                sp_qty = float(row.get("SP", 0) or 0)
                fg_qty = float(row.get("FG", 0) or 0)
                tsq = float(row.get("TSQ", 0) or 0)
                
                if tsq == 0:
                    tsq = cs_qty + gr_qty + mc_qty + sp_qty + fg_qty
                
                if tsq > 0:
                    wip_inventory[fg_code] = {
                        "CS": cs_qty,
                        "GR": gr_qty,
                        "MC": mc_qty,
                        "SP": sp_qty,
                        "FG": fg_qty,
                        "TSQ": tsq,
                    }
            
            return wip_inventory
            
        except Exception as e:
            print(f"Error loading WIP: {e}")
            return {}

# ============================================================================
# OPTIMIZER CORE
# ============================================================================

def should_skip_grinding(resource_code, cycle_time_min) -> bool:
    if pd.isna(resource_code): return True
    resource_str = str(resource_code).strip().upper()
    if resource_str in ['', '0', 'NAN', 'NONE']: return True
    if pd.isna(cycle_time_min): return True
    try:
        cycle_val = float(cycle_time_min)
        if cycle_val <= 0: return True
    except:
        return True
    return False

def build_and_solve_grinding_model(
    days: List[datetime],
    products: List[str],
    order_list: List[Dict],
    grinding_params: Dict,
    demand_units: Dict[str, float],
    wip_inventory: Dict,
    cs_from_casting: Dict[str, Dict[int, float]],
    config: GrindingConfig,
    log_path: Optional[str] = None
) -> Tuple:
    
    T = list(range(len(days)))
    max_grinding_min_per_day = config.total_grinding_capacity_per_day
    
    # Calculate CS availability
    cs_available_by_day = {}
    for j in products:
        cs_available_by_day[j] = {}
        initial_cs = wip_inventory.get(j, {}).get("CS", 0)
        for ti in T:
            cs_from_cast = cs_from_casting.get(j, {}).get(ti, 0)
            cs_available_by_day[j][ti] = initial_cs + cs_from_cast
            
    # Problem
    prob = pulp.LpProblem("GrindingSchedule", pulp.LpMinimize)
    
    # Variables
    X = pulp.LpVariable.dicts("X", (products, T), lowBound=0, cat="Integer")
    StartQty = pulp.LpVariable.dicts("StartQty", (products, T), lowBound=0, cat="Integer")
    ShortOrder = {}
    for order in order_list:
        ShortOrder[(order["part"], order["order_id"])] = pulp.LpVariable(
            f"ShortOrder_{order['part']}_{order['order_id']}", lowBound=0, cat="Integer"
        )
        
    # Helper for timing
    part_timing = {}
    for j in products:
        rc = grinding_params[j]["resource"]
        # Determine capacity for this resource
        meta = config.resource_meta.get(str(rc).strip())
        if meta:
            per_machine = float(meta["hours_per_day"]) * 60.0 * config.LINE_OEE
        else:
            per_machine = config.EFFECTIVE_DAY_MIN
            
        per_unit_min = float(grinding_params[j]["cycle_min"]) / max(1, int(grinding_params[j]["batch_qty"]))
        K = int(math.ceil(per_unit_min / per_machine))
        K = max(1, K)
        offset = K - 1
        remainder = max(0.0, per_unit_min - per_machine * (K - 1))
        contrib = [per_machine] * (K - 1) + [remainder]
        
        part_timing[j] = {
            "rc": str(rc).strip(),
            "contrib": contrib,
            "offset": offset
        }
        
    # Constraints
    
    # 1. Demand
    for j in products:
        orders_j = [o for o in order_list if o["part"] == j]
        rhs = demand_units[j]
        prob += (
            pulp.lpSum(X[j][t] for t in T) + 
            pulp.lpSum(ShortOrder[(j, o["order_id"])] for o in orders_j) == rhs,
            f"demand_{j}"
        )
        
    # 1.5 Start/Finish Link
    for j in products:
        offset = part_timing[j]["offset"]
        if offset == 0:
            for t in T:
                prob += X[j][t] == StartQty[j][t], f"sf_{j}_{t}"
        else:
            for t in T:
                if t < offset:
                    prob += X[j][t] == 0, f"no_complete_{j}_{t}"
                else:
                    prob += X[j][t] == StartQty[j][t-offset], f"spill_{j}_{t}"
        
        prob += pulp.lpSum(StartQty[j][t] for t in T) == pulp.lpSum(X[j][t] for t in T)
        
    # 2. Capacity
    parts_by_rc = defaultdict(list)
    for j in products:
        rc = str(grinding_params[j]["resource"]).strip()
        parts_by_rc[rc].append(j)
        
    rc_day_contributions = {}
    for rc, parts in parts_by_rc.items():
        rc_day_contributions[rc] = {}
        for ti in T:
            rc_day_contributions[rc][ti] = []
            for j in parts:
                contrib = part_timing[j]["contrib"]
                for d, mins in enumerate(contrib):
                    s = ti - d
                    if s >= 0 and mins > 0:
                        rc_day_contributions[rc][ti].append((j, s, mins))
                        
    for ti in T:
        for rc, parts in parts_by_rc.items():
            cap = config.capacity_by_resource_code.get(rc, max_grinding_min_per_day)
            expr = [mins * StartQty[j][s] for j, s, mins in rc_day_contributions[rc][ti]]
            prob += pulp.lpSum(expr) <= cap, f"cap_{rc}_{ti}"
            
    # 3. CS Availability
    for j in products:
        for ti in T:
            cs_avail = cs_available_by_day[j].get(ti, 0)
            if cs_avail > 0:
                prob += (
                    pulp.lpSum(StartQty[j][t] for t in T if t <= ti) <= cs_avail,
                    f"cs_{j}_{ti}"
                )
                
    # Objective
    obj_parts = []
    
    # Shortage
    obj_parts.append(config.SHORTAGE_PENALTY * pulp.lpSum(ShortOrder.values()))
    
    # Lateness (Backlog)
    for order in order_list:
        days_late = max(0, -order["days_until_due"])
        if days_late > 0:
            obj_parts.append(config.LATENESS_PENALTY_PER_DAY * days_late * ShortOrder[(order["part"], order["order_id"])])
            
    # Lead-time / Production Lateness (approximate per-part to avoid stacking)
    planning_start = days[0].date()
    for j in products:
        orders_j = [o for o in order_list if o["part"] == j]
        active_orders = [o for o in orders_j if o["due_date"].date() >= planning_start]
        if not active_orders: continue
        
        earliest_due = min(o["due_date"] for o in active_orders).date()
        
        for ti in T:
            day = days[ti].date()
            if day > earliest_due:
                diff = (day - earliest_due).days
                obj_parts.append(config.PRODUCTION_LATENESS_PENALTY * diff * X[j][ti])
            else:
                ahead = (earliest_due - day).days
                if ahead < config.LEADTIME_REQUIRED_DAYS:
                    shortfall = config.LEADTIME_REQUIRED_DAYS - ahead
                    obj_parts.append(config.LEADTIME_PENALTY_PER_DAY * shortfall * X[j][ti])
                    
    prob += pulp.lpSum(obj_parts), "Cost"
    
    # Solve
    solver_options = {
        "timeLimit": config.SOLVER_TIME_LIMIT,
        "gapRel": config.SOLVER_GAP_REL,
        "threads": config.SOLVER_THREADS
    }
    if log_path:
        solver_options["msg"] = True
        solver_options["logPath"] = log_path
    else:
        solver_options["msg"] = False
        
    solver = pulp.PULP_CBC_CMD(**solver_options)
    prob.solve(solver)
    
    if prob.status != pulp.LpStatusOptimal and prob.status != pulp.LpStatusIntegerFeasible:
        return None, None, None, None
        
    return X, StartQty, ShortOrder, cs_available_by_day

# ============================================================================
# DATA PROCESSING ORCHESTRATOR
# ============================================================================

def process_data_and_optimize(
    master_data: Dict[str, pd.DataFrame],
    casting_schedule_df: pd.DataFrame,
    start_date: datetime,
    end_date: datetime,
    config: GrindingConfig,
    log_path: Optional[str] = None
):
    # 1. Load WIP
    wip_loader = WIPLoader(master_data)
    wip_inventory = wip_loader.load_wip_inventory()
    
    # 2. Load CS from Casting
    casting_reader = CastingScheduleReader(casting_schedule_df, config)
    cs_from_casting = casting_reader.load_cs_availability(start_date, end_date)
    
    # 3. Days
    all_days = pd.date_range(start_date, end_date, freq="D")
    days = [d for d in all_days if d.weekday() < 6]
    
    # 4. Part Master & Grinding Params
    pm = master_data["Part Master"]
    grinding_params = {}
    skipped_parts = []
    
    for _, row in pm.iterrows():
        fg_code = str(row.get("FG Code", "")).strip()
        if not fg_code: continue
        
        res = row.get("Grinding Resource code")
        cycle = row.get("Grinding Cycle time (min)")
        
        if should_skip_grinding(res, cycle):
            skipped_parts.append(fg_code)
            continue
            
        grinding_params[fg_code] = {
            "cycle_min": float(cycle),
            "batch_qty": max(1, int(row.get("Grinding batch Qty", 1) or 1)),
            "resource": str(res)
        }
        
    # 5. Orders
    so = master_data["Sales Order"]
    order_list = []
    idx = 0
    
    for _, row in so.iterrows():
        fg = str(row.get("Material Code", "")).strip()
        if not fg: continue
        qty = float(row.get("Balance Qty", 0) or 0)
        if qty <= 0: continue
        
        if fg not in grinding_params: continue
        
        due = row.get("Comitted Delivery Date")
        try:
            due_date = pd.to_datetime(due)
            if pd.isna(due_date): due_date = end_date
        except:
            due_date = end_date
            
        days_until = (due_date.date() - start_date.date()).days
        try:
            due_day_idx = next(i for i, d in enumerate(days) if d.date() >= due_date.date())
        except:
            due_day_idx = len(days) - 1
            
        order_list.append({
            "part": fg,
            "order_id": idx,
            "order_no": str(row.get("Sales Order No", "")),
            "quantity": qty,
            "due_date": due_date,
            "due_day_idx": due_day_idx,
            "days_until_due": days_until
        })
        idx += 1
        
    # 6. Demand Units (Netting)
    demand_units = {}
    products = set()
    for o in order_list:
        j = o["part"]
        products.add(j)
        demand_units[j] = demand_units.get(j, 0) + o["quantity"]
        
    # Net against post-GR WIP
    for j in products:
        w = wip_inventory.get(j, {})
        done_after = float(w.get("GR", 0)) + float(w.get("MC", 0)) + float(w.get("SP", 0)) + float(w.get("FG", 0))
        demand_units[j] = max(0, demand_units[j] - done_after)
        
    products = sorted(list(products))
    
    # 7. Run Optimization
    X, StartQty, ShortOrder, cs_available = build_and_solve_grinding_model(
        days, products, order_list, grinding_params, demand_units, 
        wip_inventory, cs_from_casting, config, log_path
    )
    
    if X is None:
        return None
        
    # 8. Extract Results
    
    # Schedule
    schedule_data = []
    T = list(range(len(days)))
    
    for ti in T:
        for j in products:
            completed = X[j][ti].varValue or 0
            started = StartQty[j][ti].varValue or 0
            if completed > 0.5 or started > 0.5:
                schedule_data.append({
                    "Date": days[ti].date(),
                    "FG_Code": j,
                    "Units": round(completed),
                    "Units_Started": round(started),
                    "Resource": grinding_params[j]["resource"],
                    "CS_Available": cs_available[j].get(ti, 0)
                })
                
    schedule_df = pd.DataFrame(schedule_data)
    
    # Fulfillment
    fulfillment_data = []
    for order in order_list:
        j = order["part"]
        k = order["order_id"]
        qty = order["quantity"]
        short = ShortOrder[(j, k)].varValue or 0
        fulfilled = max(0, qty - short)
        
        status = "ON TIME"
        if short > 0:
            status = "LATE" if order["days_until_due"] < 0 else "SHORT"
            
        fulfillment_data.append({
            "FG Code": j,
            "Sales Order No": order["order_no"],
            "Due Date": order["due_date"].date(),
            "Order Qty": qty,
            "Fulfilled": fulfilled,
            "Shortage": short,
            "Status": status
        })
        
    fulfillment_df = pd.DataFrame(fulfillment_data)
    
    # Daily Capacity
    # (Simplified for return - can be expanded)
    daily_data = []
    for ti in T:
        date = days[ti].date()
        total_started = sum(StartQty[j][ti].varValue or 0 for j in products)
        daily_data.append({
            "Date": date,
            "Total_Started": total_started
        })
    daily_df = pd.DataFrame(daily_data)
    
    return {
        "schedule": schedule_df,
        "fulfillment": fulfillment_df,
        "daily": daily_df
    }


def generate_grinding_excel(results: Dict) -> io.BytesIO:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if "schedule" in results:
            results["schedule"].to_excel(writer, sheet_name="Grinding_Schedule", index=False)
        if "fulfillment" in results:
            results["fulfillment"].to_excel(writer, sheet_name="Order_Fulfillment", index=False)
        if "daily" in results:
            results["daily"].to_excel(writer, sheet_name="Daily_Summary", index=False)
            
    output.seek(0)
    return output
