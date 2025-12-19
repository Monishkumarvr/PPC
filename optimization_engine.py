"""
ORDER-AWARE CASTING SCHEDULER WITH WIP + LEAD-TIME

Refactored for Streamlit App Integration.
"""

import math
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import io

import pandas as pd
import pulp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ----------------- CONFIG CLASS -----------------

class OptimizationConfig:
    def __init__(self,
                 daily_melt_tons: float = 250.0,
                 line_hours_per_day: float = 16.0,
                 line_oee: float = 0.90,
                 shortage_penalty: float = 100000000.0,
                 lateness_penalty_per_day: float = 50.0,
                 early_production_penalty: float = 0.0,
                 leadtime_required_days: int = 14,
                 leadtime_penalty_per_day: float = 25.0,
                 production_lateness_penalty: float = 1000.0,
                 planning_date: datetime = None,
                 planning_end_date: datetime = None,
                 solver_timeout: int = 600
                 ):
        self.daily_melt_tons = daily_melt_tons
        self.daily_melt_kg = daily_melt_tons * 1000.0
        self.line_hours_per_day = line_hours_per_day
        self.line_oee = line_oee
        self.effective_day_min = line_hours_per_day * 60.0 * line_oee
        self.pattern_changeover_min = 18.0 # Keep constant or make configurable
        
        self.shortage_penalty = shortage_penalty
        self.lateness_penalty_per_day = lateness_penalty_per_day
        self.early_production_penalty = early_production_penalty
        self.leadtime_required_days = leadtime_required_days
        self.leadtime_penalty_per_day = leadtime_penalty_per_day
        self.production_lateness_penalty = production_lateness_penalty
        self.planning_date = planning_date if planning_date else datetime.now()
        self.planning_end_date = planning_end_date
        self.solver_timeout = solver_timeout

# ----------------- WIP LOADER -----------------

def load_wip_data(wip_df: Optional[pd.DataFrame]) -> Dict[str, Dict[str, float]]:
    """
    Load WIP (Work-in-Progress) inventory from Stage WIP sheet.

    Returns:
        {fg_code: {'CS': qty, 'GR': qty, 'MC': qty, 'SP': qty, 'FG': qty, 'TSQ': qty}}
    """
    try:
        if wip_df is None or wip_df.empty:
            print("\n[!] WARNING: 'Stage WIP' data not provided. Proceeding without WIP adjustment.")
            return {}

        wip_inventory: Dict[str, Dict[str, float]] = {}
        total_tsq = 0.0
        parts_with_wip = 0

        # Stats for report
        cs_count = 0
        cs1_count = 0
        cs_qty_total = 0.0
        cs1_qty_total = 0.0

        for _, row in wip_df.iterrows():
            casting_item = str(row.get("CastingItem", "")).strip()
            if not casting_item:
                continue

            # Convert CS1-XXX-XXX or CS-XXX-XXX to XXX-XXX
            # Handle CS1- first as it is longer
            if casting_item.upper().startswith("CS1-"):
                fg_code = casting_item[4:]
                cs1_count += 1
            elif casting_item.upper().startswith("CS-"):
                fg_code = casting_item[3:]
                cs_count += 1
            else:
                fg_code = casting_item

            cs_qty = float(row.get("CS", 0) or 0)
            gr_qty = float(row.get("GR", 0) or 0)
            mc_qty = float(row.get("MC", 0) or 0)
            sp_qty = float(row.get("SP", 0) or 0)
            fg_qty = float(row.get("FG", 0) or 0)
            tsq = float(row.get("TSQ", 0) or 0)

            if tsq == 0:
                tsq = cs_qty + gr_qty + mc_qty + sp_qty + fg_qty

            if tsq > 0:
                wip_inventory[fg_code] = {
                    "CS": cs_qty,
                    "GR": gr_qty,
                    "MC": mc_qty,
                    "SP": sp_qty,
                    "FG": fg_qty,
                    "TSQ": tsq,
                }
                total_tsq += tsq
                parts_with_wip += 1
                
                if casting_item.upper().startswith("CS1-"):
                    cs1_qty_total += tsq
                elif casting_item.upper().startswith("CS-"):
                    cs_qty_total += tsq

        print(f"✓ WIP Data Loaded:")
        print(f"  Parts with WIP: {parts_with_wip}")
        print(f"  Total TSQ: {total_tsq} units")
        print(f"  Prefix breakdown:")
        print(f"    CS- prefix:   {cs_count} items,   {cs_qty_total} units")
        print(f"    CS1- prefix:  {cs1_count} items,  {cs1_qty_total} units")

        return wip_inventory

    except Exception as e:
        print(f"\n[!] ERROR loading WIP data: {e}")
        return {}


# ----------------- DATA LOADER (WIP + ORDERS) -----------------

def process_casting_data(
    input_data: Any,
    config: OptimizationConfig,
):
    """
    Load all casting data from either an uploaded file or a dictionary of DataFrames.
    
    Args:
        input_data: Either bytes/file object (Excel) OR Dict[str, pd.DataFrame]
        config: OptimizationConfig
    """
    
    dfs = {}
    if isinstance(input_data, dict):
        # Assume it's already a dict of DataFrames
        dfs = input_data
    else:
        # Assume it's a file path or file-like object
        dfs = pd.read_excel(input_data, sheet_name=None)

    part_master = dfs.get("Part Master")
    sales_order = dfs.get("Sales Order")
    machine_constraints = dfs.get("Machine Constraints")
    mould_box_capacity = dfs.get("Mould Box Capacity")
    
    if part_master is None or sales_order is None or machine_constraints is None or mould_box_capacity is None:
        raise ValueError("Missing required sheets in input data.")

    # Determine date range
    sales_order["Delivery Date"] = pd.to_datetime(
        sales_order["Comitted Delivery Date"], errors="coerce"
    )
    # Filter out NaT dates if any to avoid errors in max/min
    valid_dates = sales_order["Delivery Date"].dropna()
    
    start_date = config.planning_date
    
    # --- INTELLIGENT HORIZON SELECTION ---
    if config.planning_end_date:
        end_date = config.planning_end_date
    elif not valid_dates.empty:
        max_date = valid_dates.max()
        # CAP the auto-horizon to 90 days to prevent crashes due to massive horizons (e.g. 1 year)
        limit_date = start_date + pd.Timedelta(days=90)
        
        if max_date > limit_date:
            print(f"Warning: Max order date is {max_date}, but limiting planning horizon to 90 days ({limit_date}) for performance. Please specify an end date to override.")
            end_date = limit_date
        else:
            end_date = max_date
    else:
        # Fallback if no dates and no end date specified
        end_date = config.planning_date + pd.Timedelta(days=30)

    # WIP inventory (may be empty)
    wip_inventory = load_wip_data(dfs.get("Stage WIP"))

    # Working days (Mon–Sat)
    all_days = pd.date_range(start_date, end_date, freq="D")
    days = [d for d in all_days if d.weekday() < 6]
    
    if not days: # Fallback if start > end or no working days
         days = [start_date]

    # Part master lookup
    pm_by_fg = {
        str(row["FG Code"]).strip(): row
        for _, row in part_master.iterrows()
    }

    # Product-level structures
    bunch_weight_kg: Dict[str, float] = {}
    box_qty: Dict[str, float] = {}
    line_time_min: Dict[str, float] = {}
    cycle_days: Dict[str, int] = {}
    line: Dict[str, str] = {}
    box_size_of: Dict[str, str] = {}

    valid_products: List[str] = []
    skipped_products = []

    # Unique products from sales order
    all_products = sales_order["Material Code"].unique()

    for fg_raw in all_products:
        fg = str(fg_raw).strip()

        # Skip pattern orders
        if fg.upper().startswith("PAT"):
            total_qty = sales_order[sales_order["Material Code"] == fg]["Balance Qty"].sum()
            skipped_products.append((fg, total_qty, "Pattern order (not a production order)"))
            continue

        if fg not in pm_by_fg:
            total_qty = sales_order[sales_order["Material Code"] == fg]["Balance Qty"].sum()
            skipped_products.append((fg, total_qty, "Not in Part Master"))
            continue

        row = pm_by_fg[fg]

        # Box quantity
        bq = float(row.get("Box Quantity", 0) or 0)
        if bq <= 0:
            bq = float(row.get("Casting Batch Qty", 1) or 1)
        if bq <= 0:
            total_qty = sales_order[sales_order["Material Code"] == fg]["Balance Qty"].sum()
            skipped_products.append((fg, total_qty, "Invalid box quantity"))
            continue
        box_qty[fg] = bq

        # Bunch weight
        bunch = float(row.get("Bunch Wt.", 0) or 0)
        if bunch <= 0:
            std_wt = float(row.get("Standard unit wt.", 0) or 0)
            casting_batch = float(row.get("Casting Batch Qty", 1) or 1)
            bunch = casting_batch * std_wt * 1.2
        bunch_weight_kg[fg] = bunch

        # Times
        casting_cycle_min = float(row.get("Casting Cycle time (min)", 0) or 0)
        vacuum_time_min = float(row.get("Vacuum Time (min)", 0) or 0)
        cooling_hrs = float(row.get("Cooling Time (hrs)", 0) or 0)
        shakeout_min = float(row.get("Shakeout Time (min)", 0) or 0)

        line_time_min[fg] = vacuum_time_min + casting_cycle_min

        full_cycle_hrs = (vacuum_time_min + casting_cycle_min + shakeout_min) / 60.0 + cooling_hrs
        cycle_days[fg] = max(1, math.floor(full_cycle_hrs / 24.0))

        # Line
        ml = str(row.get("Moulding Line", "")).lower()
        if "small" in ml:
            line[fg] = "small"
        elif "big" in ml:
            line[fg] = "big"
        else:
            line[fg] = "small"

        # Box size
        box_size_of[fg] = str(row.get("Box Size", "")).strip()

        valid_products.append(fg)

    products = valid_products

    # Physical boxes
    box_max_boxes: Dict[str, float] = {}
    for _, row in mould_box_capacity.iterrows():
        size = str(row["Box_Size"]).strip()
        if not size:
            continue
        num_boxes = float(row.get("No_of_boxes", 0) or 0)
        if num_boxes > 0:
            box_max_boxes[size] = num_boxes

    # Capacity
    max_melt_kg_per_day = config.daily_melt_kg

    small_row = machine_constraints[
        machine_constraints["Resource Name"].astype(str).str.contains("Small Vacuum", case=False, na=False)
    ].iloc[0]
    big_row = machine_constraints[
        machine_constraints["Resource Name"].astype(str).str.contains("Big Vacuum", case=False, na=False)
    ].iloc[0]

    def line_effective_minutes(row) -> float:
        avail = float(row.get("Available Hours per Day", 0) or 0)
        shifts = float(row.get("No of Shift", 1) or 1)
        machines = float(row.get("No Of Resource", 1) or 1)
        hours_total = min(avail * shifts, config.line_hours_per_day)
        return hours_total * 60.0 * config.line_oee * machines

    max_time_small_min = line_effective_minutes(small_row)
    max_time_big_min = line_effective_minutes(big_row)

    # ---- ORDER LIST (same structure as wip.py) ----

    order_list = []
    order_counter = 0

    for j in products:
        orders_j = sales_order[sales_order["Material Code"] == j].copy()
        orders_j = orders_j.sort_values("Delivery Date")

        for _, order_row in orders_j.iterrows():
            due_date = order_row["Delivery Date"]
            qty_pieces = order_row["Balance Qty"]

            if qty_pieces <= 0:
                continue

            bq = box_qty.get(j, 1)
            if bq <= 0:
                bq = 1
            qty_boxes = math.ceil(qty_pieces / bq)

            # find due_day_idx in working days
            due_day_idx = None
            if pd.notna(due_date):
                for ti, day in enumerate(days):
                    if day.date() >= due_date.date():
                        due_day_idx = ti
                        break
            if due_day_idx is None:
                due_day_idx = len(days) - 1

            if pd.notna(due_date):
                days_until_due = (due_date.date() - days[0].date()).days
            else:
                days_until_due = 1000

            order_list.append({
                "part": j,
                "order_id": order_counter,
                "due_day_idx": due_day_idx,
                "qty_boxes": qty_boxes,
                "qty_pieces": qty_pieces,
                "due_date": due_date if pd.notna(due_date) else datetime(2099, 12, 31),
                "days_until_due": days_until_due,
                "sales_order_no": order_row.get("Sales Order No", "Unknown"),
            })
            order_counter += 1

    # ---- Gross demand (boxes) per FG from orders ----

    demand_boxes: Dict[str, float] = {}
    gross_demand_boxes: Dict[str, float] = {}
    wip_coverage_boxes: Dict[str, float] = {}

    for j in products:
        orders_j = [o for o in order_list if o["part"] == j]
        gross_boxes = sum(o["qty_boxes"] for o in orders_j)
        gross_demand_boxes[j] = gross_boxes

        # Apply WIP (TSQ) reduction if available
        if j in wip_inventory:
            tsq_units = wip_inventory[j]["TSQ"]
            bq = box_qty.get(j, 1)
            tsq_boxes = math.floor(tsq_units / bq)
        else:
            tsq_boxes = 0

        net_boxes = max(0, gross_boxes - tsq_boxes)
        wip_coverage_boxes[j] = tsq_boxes
        demand_boxes[j] = net_boxes

    return (
        products,
        days,
        demand_boxes,
        bunch_weight_kg,
        box_qty,
        line_time_min,
        cycle_days,
        line,
        box_size_of,
        box_max_boxes,
        max_melt_kg_per_day,
        max_time_small_min,
        max_time_big_min,
        order_list,
        wip_coverage_boxes,
        gross_demand_boxes,
    )


# ----------------- VALIDATION & RECOMMENDATION -----------------

def validate_excel_sheets(input_data: Any) -> Dict[str, bool]:
    """
    Validate the presence of required sheets in the Excel file or DataFrame dict.
    Returns a dict {sheet_name: is_present}
    """
    required_sheets = ["Part Master", "Sales Order", "Machine Constraints", "Mould Box Capacity"]
    optional_sheets = ["Stage WIP"]
    
    try:
        sheet_names = []
        if isinstance(input_data, dict):
            sheet_names = list(input_data.keys())
        else:
            xls = pd.ExcelFile(input_data)
            sheet_names = xls.sheet_names
        
        status = {}
        for sheet in required_sheets + optional_sheets:
            status[sheet] = sheet in sheet_names
            
        return status
    except Exception:
        return {sheet: False for sheet in required_sheets + optional_sheets}


def generate_recommendations(
    capacity_rows: List[Dict], 
    box_utilization_rows: List[Dict], 
    shortage_rows: List[Dict]
) -> List[Dict]:
    """
    Generate heuristic recommendations based on optimization results.
    Returns a list of dicts with 'type', 'message', 'severity'.
    """
    recommendations = []
    
    # 1. Analyze Order Shortages
    df_shortage = pd.DataFrame(shortage_rows)
    if not df_shortage.empty:
        total_shortage_orders = len(df_shortage[df_shortage['Status'] != 'ON TIME'])
        late_orders = len(df_shortage[df_shortage['Status'] == 'LATE'])
        
        if total_shortage_orders > 0:
            recommendations.append({
                "type": "Shortage",
                "message": f"{total_shortage_orders} orders are delayed or short. {late_orders} are already past due date. Consider expediting production for high-priority late orders.",
                "severity": "High"
            })
            
            # Identify most delayed product
            if 'Days Until Due' in df_shortage.columns:
                worst_order = df_shortage.sort_values('Days Until Due').iloc[0]
                if worst_order['Days Until Due'] < 0:
                    recommendations.append({
                        "type": "Prioritization",
                        "message": f"Critical Delay: Order '{worst_order['Sales Order No']}' for '{worst_order['FG Code']}' is overdue by {abs(worst_order['Days Until Due'])} days. Prioritize this FG Code on the line immediately.",
                        "severity": "Critical"
                    })

    # 2. Analyze Box Bottlenecks
    df_boxes = pd.DataFrame(box_utilization_rows)
    if not df_boxes.empty:
        # Group by box size and calculate average utilization
        box_summary = df_boxes.groupby('Box_Size').agg({
            'Utilization_%': 'mean',
            'Max_Boxes': 'first'
        }).reset_index()
        
        critical_boxes = box_summary[box_summary['Utilization_%'] > 85]
        
        for _, row in critical_boxes.iterrows():
            box_size = row['Box_Size']
            util = row['Utilization_%']
            current_inv = row['Max_Boxes']
            
            # Simple heuristic: suggest adding 10-20% more boxes
            suggested_add = max(2, int(current_inv * 0.15))
            
            recommendations.append({
                "type": "Capacity",
                "message": f"High usage detected for Box Size {box_size} (Avg Util: {util:.1f}%). This is likely a bottleneck. Recommendation: Purchase {suggested_add} additional boxes of this size to relieve pressure.",
                "severity": "High" if util > 95 else "Medium"
            })

    # 3. Analyze Line/Melt Capacity
    df_cap = pd.DataFrame(capacity_rows)
    if not df_cap.empty:
        avg_melt = df_cap['Melt_Utilization_%'].mean()
        max_melt = df_cap['Melt_Utilization_%'].max()
        
        avg_small = df_cap['Small_Utilization_%'].mean()
        avg_big = df_cap['Big_Utilization_%'].mean()
        
        if avg_melt > 90:
             recommendations.append({
                "type": "Capacity",
                "message": f"Melt capacity is critically tight (Avg: {avg_melt:.1f}%, Max: {max_melt:.1f}%). Moulding is likely limited by liquid metal availability. Consider adding a shift or increasing melt batch frequency.",
                "severity": "High"
            })
        
        if avg_small > 90:
             recommendations.append({
                "type": "Capacity",
                "message": f"Small Vacuum Line is heavily loaded (Avg: {avg_small:.1f}%). Check if some parts can be moved to Big Line or if cycle times can be optimized.",
                "severity": "Medium"
            })
            
        if avg_big > 90:
             recommendations.append({
                "type": "Capacity",
                "message": f"Big Vacuum Line is heavily loaded (Avg: {avg_big:.1f}%). Ensure OEE is maximized during peak days.",
                "severity": "Medium"
            })

    # 4. WIP Analysis
    # If we had access to input wip stats easily here we could add it, but based on results:
    # (We don't pass raw WIP stats to this function, only outputs, so skip for now)
    
    if not recommendations:
        recommendations.append({
            "type": "General",
            "message": "Production plan looks healthy. No critical resource bottlenecks or major shortages detected.",
            "severity": "Low"
        })
        
    return recommendations


# ----------------- MILP SOLVER (WIP + LEAD-TIME) -----------------

def build_and_solve_enhanced_milp(
    products: List[str],
    days: List[pd.Timestamp],
    demand_boxes: Dict[str, float],
    bunch_weight_kg: Dict[str, float],
    box_qty: Dict[str, float],
    line_time_min: Dict[str, float],
    cycle_days: Dict[str, int],
    line: Dict[str, str],
    box_size_of: Dict[str, str],
    box_max_boxes: Dict[str, float],
    max_melt_kg_per_day: float,
    max_time_small_min: float,
    max_time_big_min: float,
    order_list: List[Dict],
    wip_coverage_boxes: Dict[str, float],
    gross_demand_boxes: Dict[str, float],
    config: OptimizationConfig,
    log_path: Optional[str] = None
):
    T = list(range(len(days)))
    
    # Debug/Log problem size
    num_products = len(products)
    num_days = len(days)
    num_orders = len(order_list)
    print(f"INFO: Building Model. Products: {num_products}, Days: {num_days}, Orders: {num_orders}")
    
    small_products = [j for j in products if line.get(j, "small") == "small"]
    big_products = [j for j in products if line.get(j, "small") == "big"]

    prob = pulp.LpProblem("EnhancedCastingPlan_WIP", pulp.LpMinimize)

    # Variables
    X = pulp.LpVariable.dicts("X", (products, T), lowBound=0, cat="Integer")
    
    # OPTIMIZATION: Only create Y variables if changeover cost > 0
    use_setup_vars = config.pattern_changeover_min > 0.1
    Y = {}
    if use_setup_vars:
        Y = pulp.LpVariable.dicts("Y", (products, T), lowBound=0, upBound=1, cat="Binary")

    ShortOrder: Dict[Tuple[str, int], pulp.LpVariable] = {}
    for order in order_list:
        j = order["part"]
        k = order["order_id"]
        ShortOrder[(j, k)] = pulp.LpVariable(
            f"ShortOrder_{j}_{k}", lowBound=0, cat="Integer"
        )

    # Constraints

    # 1. Demand balance (net demand after WIP)
    for j in products:
        orders_j = [o for o in order_list if o["part"] == j]
        rhs = demand_boxes[j]  # already gross - TSQ

        prob += (
            pulp.lpSum(X[j][t] for t in T)
            + pulp.lpSum(ShortOrder[(j, o["order_id"])] for o in orders_j)
            == rhs,
            f"demand_{j}",
        )

    # 2. Capacity constraints (melt + line + boxes)
    melt_today: Dict[int, pulp.LpAffineExpression] = {}
    small_time_today: Dict[int, pulp.LpAffineExpression] = {}
    big_time_today: Dict[int, pulp.LpAffineExpression] = {}

    for ti in T:
        # Melt
        melt_today[ti] = pulp.lpSum(
            bunch_weight_kg.get(j, 0) * X[j][ti] for j in products
        )
        prob += melt_today[ti] <= max_melt_kg_per_day, f"melt_cap_{ti}"

        # Small line
        if small_products:
            production_time = pulp.lpSum(line_time_min.get(j, 0) * X[j][ti] for j in small_products)
            setup_time = 0
            if use_setup_vars:
                setup_time = config.pattern_changeover_min * pulp.lpSum(Y[j][ti] for j in small_products)
                
            small_time_today[ti] = production_time + setup_time
            prob += small_time_today[ti] <= max_time_small_min, f"small_cap_{ti}"
        else:
            small_time_today[ti] = 0

        # Big line
        if big_products:
            production_time = pulp.lpSum(line_time_min.get(j, 0) * X[j][ti] for j in big_products)
            setup_time = 0
            if use_setup_vars:
                setup_time = config.pattern_changeover_min * pulp.lpSum(Y[j][ti] for j in big_products)
            
            big_time_today[ti] = production_time + setup_time
            prob += big_time_today[ti] <= max_time_big_min, f"big_cap_{ti}"
        else:
            big_time_today[ti] = 0

        # Box capacity (daily reset, like wip.py)
        for box_size, n_boxes in box_max_boxes.items():
            js_for_size = [j for j in products if box_size_of.get(j) == box_size]
            if not js_for_size:
                continue
            prob += (
                pulp.lpSum(X[j][ti] for j in js_for_size) <= n_boxes,
                f"box_daily_{box_size}_{ti}",
            )

    # 3. Link X & Y (ONLY IF SETUP VARS ARE USED)
    if use_setup_vars:
        BIG_M = {j: max(demand_boxes.get(j, 0), 1) for j in products}
        for j in products:
            for ti in T:
                prob += X[j][ti] <= BIG_M[j] * Y[j][ti], f"link_{j}_{ti}"

    # Objective
    obj_parts = []

    # Part 1: Shortage penalty (per order, on net demand after WIP)
    shortage_expr = pulp.lpSum(ShortOrder[(o["part"], o["order_id"])] for o in order_list)
    obj_parts.append(config.shortage_penalty * shortage_expr)

    # Part 2: Lateness penalty on backlog orders (as in wip.py)
    for order in order_list:
        j = order["part"]
        k = order["order_id"]
        days_until_due = order["days_until_due"]
        days_late = max(0, -days_until_due)
        if days_late > 0:
            obj_parts.append(config.lateness_penalty_per_day * days_late * ShortOrder[(j, k)])

    # Part 3: Production timing penalty (production AFTER due date)
    for order in order_list:
        j = order["part"]
        due_idx = min(order["due_day_idx"], len(T) - 1)
        for ti in T:
            if ti > due_idx:
                days_late = ti - due_idx
                obj_parts.append(config.production_lateness_penalty * days_late * X[j][ti])

    # Part 4: Lead-time penalty (production TOO CLOSE to due date)
    # For each order and each day <= due_date:
    #   if (due_date - day) < LEADTIME_REQUIRED_DAYS, penalize X for being "late start"
    for order in order_list:
        j = order["part"]
        due_date = order["due_date"].date()
        for ti in T:
            day = days[ti].date()
            if day > due_date:
                continue  # already covered by production-late term
            days_ahead = (due_date - day).days
            if days_ahead < config.leadtime_required_days:
                shortfall = config.leadtime_required_days - days_ahead
                obj_parts.append(config.leadtime_penalty_per_day * shortfall * X[j][ti])

    # Part 5: Early production penalty (optional)
    if config.early_production_penalty > 0:
        for order in order_list:
            j = order["part"]
            due_idx = min(order["due_day_idx"], len(T) - 1)
            for ti in T:
                if ti < due_idx:
                    days_early = due_idx - ti
                    obj_parts.append(config.early_production_penalty * days_early * X[j][ti])

    # Part 6: Pattern changeover penalties
    if use_setup_vars:
        for ti in T:
            for j in products:
                obj_parts.append(config.pattern_changeover_min * Y[j][ti])

    prob += pulp.lpSum(obj_parts), "TotalCost"

    # Solve
    # If log_path is provided, we direct output there and set msg=True
    solver_options = {
        "timeLimit": config.solver_timeout,
        "gapRel": 0.005,
        "threads": 4, # Reduce threads to be safe
    }
    
    if log_path:
        solver_options["msg"] = True
        solver_options["logPath"] = log_path
    else:
        solver_options["msg"] = False

    solver = pulp.PULP_CBC_CMD(**solver_options)
    status = prob.solve(solver)
    
    # Check solver status
    print(f"\n{'='*80}")
    print(f"SOLVER STATUS: {pulp.LpStatus[status]}")
    print(f"{'='*80}\n")
    
    if status == pulp.LpStatusInfeasible:
        print("❌ PROBLEM IS INFEASIBLE")
        print("   No feasible solution exists with current constraints")
        print(f"   Problem size: {prob.numVariables()} variables, {prob.numConstraints()} constraints")
        return prob, [], [], [], []
    elif status != pulp.LpStatusOptimal and status != pulp.LpStatusIntegerFeasible:
        print(f"⚠️  Solver status: {pulp.LpStatus[status]}")
        return prob, [], [], [], []
    
    print("✅ Optimal solution found!\n")

    if prob.status != pulp.LpStatusOptimal and prob.status != pulp.LpStatusIntegerFeasible:
        # Just return empty if failed, or handle gracefully
        # But we will try to extract whatever we can even if not optimal
        pass
        
    # Extract schedules & reports
    schedule_rows: List[Dict] = []
    daily_capacity_rows: List[Dict] = []
    box_utilization_rows: List[Dict] = []

    for ti, day in enumerate(days):
        daily_melt = 0.0
        daily_small_time = 0.0
        daily_big_time = 0.0
        daily_small_boxes = 0
        daily_big_boxes = 0

        for j in products:
            x_jt = X[j][ti].value()
            if x_jt is None or x_jt <= 1e-6:
                continue

            boxes = int(round(x_jt))
            if boxes <= 0:
                continue

            schedule_rows.append({
                "Date": day.date(),
                "FG Code": j,
                "Line": line.get(j, "small"),
                "Box_Size": box_size_of.get(j, "Unknown"),
                "Boxes": boxes,
                "Units": boxes * box_qty.get(j, 1),
                "Melt_kg": boxes * bunch_weight_kg.get(j, 0),
            })

            daily_melt += boxes * bunch_weight_kg.get(j, 0)

            if line.get(j, "small") == "small":
                daily_small_boxes += boxes
                daily_small_time += boxes * line_time_min.get(j, 0)
            else:
                daily_big_boxes += boxes
                daily_big_time += boxes * line_time_min.get(j, 0)

        small_products_today = 0
        big_products_today = 0
        
        if use_setup_vars:
             small_products_today = len([j for j in products if X[j][ti].value() and X[j][ti].value() > 0.1 and line.get(j) == "small"])
             big_products_today = len([j for j in products if X[j][ti].value() and X[j][ti].value() > 0.1 and line.get(j) == "big"])

        daily_small_time += config.pattern_changeover_min * small_products_today
        daily_big_time += config.pattern_changeover_min * big_products_today

        daily_capacity_rows.append({
            "Date": day.date(),
            "Melt_Used_kg": daily_melt,
            "Melt_Capacity_kg": max_melt_kg_per_day,
            "Melt_Utilization_%": (daily_melt / max_melt_kg_per_day * 100) if max_melt_kg_per_day > 0 else 0,
            "Small_Time_Used_min": daily_small_time,
            "Small_Time_Capacity_min": max_time_small_min,
            "Small_Utilization_%": (daily_small_time / max_time_small_min * 100) if max_time_small_min > 0 else 0,
            "Small_Boxes": daily_small_boxes,
            "Big_Time_Used_min": daily_big_time,
            "Big_Time_Capacity_min": max_time_big_min,
            "Big_Utilization_%": (daily_big_time / max_time_big_min * 100) if max_time_big_min > 0 else 0,
            "Big_Boxes": daily_big_boxes,
        })

        # Multi-day box occupancy using cycle_days (same as wip.py)
        for box_size, max_boxes in box_max_boxes.items():
            js_for_size = [j for j in products if box_size_of.get(j) == box_size]
            if not js_for_size:
                continue

            cast_today = 0
            in_process = 0

            max_turnaround = max([cycle_days.get(j, 2) for j in js_for_size])

            for lookback in range(max_turnaround + 1):
                if ti - lookback < 0:
                    continue
                for j in js_for_size:
                    x_val = X[j][ti - lookback].value()
                    if x_val and x_val > 0.1:
                        boxes = int(round(x_val))
                        product_turnaround = cycle_days.get(j, 2)
                        if lookback < product_turnaround:
                            if lookback == 0:
                                cast_today += boxes
                            else:
                                in_process += boxes

            total_occupied = cast_today + in_process
            available_today = max_boxes - total_occupied
            utilization_pct = (total_occupied / max_boxes * 100) if max_boxes > 0 else 0

            if available_today < 0:
                status = "VIOLATION"
            elif available_today == 0:
                status = "MAXED"
            elif available_today < max_boxes * 0.2:
                status = "BUSY"
            else:
                status = "AVAILABLE"

            box_utilization_rows.append({
                "Date": day.date(),
                "Box_Size": box_size,
                "Max_Boxes": max_boxes,
                "Cast_Today": cast_today,
                "In_Process": in_process,
                "Total_Occupied": total_occupied,
                "Available_Today": available_today,
                "Utilization_%": round(utilization_pct, 1),
                "Status": status,
            })

    # Order-level shortage summary (unchanged, but now vs net demand)
    # ---------------------------------------------------------
    # FIFO ALLOCATION FOR ORDER-LEVEL DATES
    # ---------------------------------------------------------
    # Map each order to its production days to find First/Last Moulding Date
    
    # 1. Gather all production per part: {part: [(day_idx, date, boxes_produced), ...]}
    production_by_part = {j: [] for j in products}
    for j in products:
        for ti in T:
            val = X[j][ti].value()
            if val and val > 0.1:
                boxes = int(round(val))
                production_by_part[j].append((ti, days[ti].date(), boxes))
        # Sort by day just in case
        production_by_part[j].sort(key=lambda x: x[0])

    # 2. Iterate orders (sorted by due date) and allocate
    order_dates = {} # order_id -> {first_date, last_date}
    
    # Work on a copy of production to consume it
    remaining_production = {j: [list(item) for item in production_by_part[j]] for j in products}
    
    # Sort orders by due date to respect FIFO fulfillment assumption
    # (Though the optimization might have shuffled priorities, FIFO is standard for "when did my order get made")
    sorted_orders = sorted(order_list, key=lambda o: o["due_date"])
    
    for order in sorted_orders:
        j = order["part"]
        oid = order["order_id"]
        qty_needed = order["qty_boxes"]
        
        # Calculate shortage to know how much was actually made for this order
        short_val = ShortOrder[(j, oid)].value() or 0
        qty_fulfilled = max(0, qty_needed - int(round(short_val)))
        
        if qty_fulfilled <= 0:
            order_dates[oid] = {"first": None, "last": None}
            continue
            
        first_date = None
        last_date = None
        
        # Consume from remaining production
        prod_stream = remaining_production[j]
        
        while qty_fulfilled > 0 and prod_stream:
            day_idx, day_date, avail = prod_stream[0]
            
            if avail <= 0:
                prod_stream.pop(0)
                continue
                
            take = min(qty_fulfilled, avail)
            
            if first_date is None:
                first_date = day_date
            last_date = day_date
            
            # Update stream
            prod_stream[0][2] -= take
            qty_fulfilled -= take
            
            if prod_stream[0][2] <= 0:
                prod_stream.pop(0)
                
        order_dates[oid] = {"first": first_date, "last": last_date}

    # ---------------------------------------------------------
    # GENERATE REPORT
    # ---------------------------------------------------------

    order_shortage_rows: List[Dict] = []

    for order in order_list:
        j = order["part"]
        k = order["order_id"]

        short = ShortOrder[(j, k)].value() or 0
        short_boxes = int(round(short))
        short_pieces = short_boxes * box_qty.get(j, 1)

        days_late = max(0, -order["days_until_due"])

        if short_boxes == 0:
            status = "ON TIME"
        elif order["days_until_due"] < 0:
            status = "LATE"
        else:
            status = "SHORT"
            
        # Extended fields
        dates = order_dates.get(k, {"first": None, "last": None})
        first_mould_date = dates["first"]
        last_mould_date = dates["last"]
        
        days_before_delivery = None
        lead_time_status = "N/A"
        lead_time_penalty_val = 0.0
        
        if first_mould_date:
            due_date_date = order["due_date"].date()
            delta = (due_date_date - first_mould_date).days
            days_before_delivery = delta
            
            if delta < config.leadtime_required_days:
                lead_time_status = "VIOLATION"
                shortfall = config.leadtime_required_days - delta
                # Approximate penalty calculation for reporting (cost is abstract)
                lead_time_penalty_val = shortfall * config.leadtime_penalty_per_day * (order["qty_boxes"] - short_boxes)
            else:
                lead_time_status = "OK"

        order_shortage_rows.append({
            "FG Code": j,
            "Sales Order No": order["sales_order_no"],
            "Due Date": order["due_date"].date(),
            "First Moulding Date": first_mould_date,
            "Last Moulding Date": last_mould_date,
            "Days Before Delivery": days_before_delivery,
            "Lead Time Status": lead_time_status,
            "Lead Time Penalty": lead_time_penalty_val,
            "Days Until Due": order["days_until_due"],
            "Order Qty (pieces)": order["qty_pieces"],
            "Order Qty (boxes)": order["qty_boxes"],
            "Shortage (boxes)": short_boxes,
            "Shortage (pieces)": short_pieces,
            "Fulfillment %": 0.0 if short_boxes >= order["qty_boxes"] else round((1 - short_boxes/order["qty_boxes"])*100, 1),
            "Status": status,
            "Priority": "HIGH" if days_late > 0 else ("MEDIUM" if order["days_until_due"] < 30 else "LOW"),
        })

    return prob, schedule_rows, order_shortage_rows, daily_capacity_rows, box_utilization_rows


def generate_excel_output(schedule_rows, order_shortage_rows, daily_capacity_rows, box_utilization_rows):
    output = io.BytesIO()
    
    def format_dates_as_strings(df: pd.DataFrame) -> pd.DataFrame:
        for col in df.columns:
            if df[col].dtype == "object" or "date" in col.lower():
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                    if pd.api.types.is_datetime64_any_dtype(df[col]):
                        df[col] = df[col].dt.strftime("%Y-%m-%d")
                except Exception:
                    pass
        return df

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if schedule_rows:
            df_schedule = pd.DataFrame(schedule_rows).sort_values(["Date", "FG Code"])
            df_schedule = format_dates_as_strings(df_schedule)
            df_schedule.to_excel(writer, sheet_name="Production_Schedule", index=False)

        if order_shortage_rows:
            df_shortage = pd.DataFrame(order_shortage_rows).sort_values(
                ["Status", "Days Until Due", "FG Code"]
            )
            df_shortage = format_dates_as_strings(df_shortage)
            df_shortage.to_excel(writer, sheet_name="Order_Fulfillment", index=False)

        if daily_capacity_rows:
            df_capacity = pd.DataFrame(daily_capacity_rows)
            df_capacity = format_dates_as_strings(df_capacity)
            df_capacity.to_excel(writer, sheet_name="Daily_Capacity", index=False)

        if box_utilization_rows:
            df_boxes = pd.DataFrame(box_utilization_rows)
            df_boxes = format_dates_as_strings(df_boxes)
            df_boxes.to_excel(writer, sheet_name="Daily_Box_Status", index=False)

            # Summary sheet
            summary_rows = []
            for box_size in df_boxes["Box_Size"].unique():
                box_data = df_boxes[df_boxes["Box_Size"] == box_size]
                if box_data.empty:
                    continue
                total_days = len(box_data)
                avg_util = box_data["Utilization_%"].mean()
                max_util = box_data["Utilization_%"].max()
                bottleneck_days = len(
                    box_data[box_data["Status"].isin(["MAXED", "VIOLATION"])]
                )
                bottleneck_pct = (bottleneck_days / total_days * 100) if total_days > 0 else 0

                summary_rows.append({
                    "Box_Size": box_size,
                    "Max_Boxes": box_data["Max_Boxes"].iloc[0],
                    "Total_Days": total_days,
                    "Avg_Utilization_%": round(avg_util, 1),
                    "Max_Utilization_%": round(max_util, 1),
                    "Bottleneck_Days": bottleneck_days,
                    "Bottleneck_%": round(bottleneck_pct, 1),
                })

            df_summary = pd.DataFrame(summary_rows)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)

            # Investment plan
            investment_rows = []
            for _, row in df_summary.iterrows():
                bottleneck_pct = row["Bottleneck_%"]
                if bottleneck_pct > 95:
                    priority = "CRITICAL"
                    additional_boxes = max(5, int(row["Max_Boxes"] * 0.2))
                elif bottleneck_pct > 80:
                    priority = "HIGH"
                    additional_boxes = max(3, int(row["Max_Boxes"] * 0.15))
                elif bottleneck_pct > 50:
                    priority = "MEDIUM"
                    additional_boxes = max(2, int(row["Max_Boxes"] * 0.10))
                else:
                    priority = "LOW"
                    additional_boxes = 0

                recommendation = (
                    f"Add {additional_boxes} more boxes"
                    if additional_boxes > 0
                    else "No action needed"
                )

                investment_rows.append({
                    "Box_Size": row["Box_Size"],
                    "Bottleneck_%": row["Bottleneck_%"],
                    "Priority": priority,
                    "Recommendation": recommendation,
                    "Additional_Boxes_Needed": additional_boxes,
                })

            df_investment = pd.DataFrame(investment_rows)
            if not df_investment.empty:
                df_investment = df_investment.sort_values(
                    "Bottleneck_%", ascending=False
                )
            df_investment.to_excel(writer, sheet_name="Investment_Plan", index=False)
    
    # Apply coloring (requires saving and reloading, but since we are in memory...)
    # OpenPyXL cannot easily load from BytesIO if we just wrote to it without seeking.
    # But pd.ExcelWriter saves it.
    
    output.seek(0)
    
    # We can skip the complex coloring for now or implement it if critical.
    # The original script used openpyxl to color. 
    # To do that in memory:
    
    wb = load_workbook(output)
    if "Daily_Box_Status" in wb.sheetnames:
        ws = wb["Daily_Box_Status"]
        status_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "Status":
                status_col = col_idx
                break

        if status_col:
            for row_idx in range(2, ws.max_row + 1):
                status_cell = ws.cell(row=row_idx, column=status_col)
                status_value = status_cell.value

                if status_value in ["VIOLATION", "MAXED"]:
                    fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif status_value == "BUSY":
                    fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                elif status_value == "AVAILABLE":
                    fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                else:
                    continue

                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
                    
         # Auto width logic
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws[col_letter]:
                if row.value:
                    try:
                        max_length = max(max_length, len(str(row.value)))
                    except:
                        pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output
